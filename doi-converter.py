from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException


import sys
import re
import openpyxl
import time

def find_col(sheet):
  #if the word 'doi' is in the title of the column
  for i in range(1, sheet.max_column + 1):
      if 'doi' in (sheet.cell(row = 1, column = i).value).lower():
        col_val = sheet.cell(row = 2, column = i).value
        if col_val.find("10.") != -1:
            return i
  return -1


def read_col():
    while True:
        try: 
            fname = input("Enter the full input file name (including the extension): " )
            workbook = openpyxl.load_workbook(filename=fname)
            sheet = workbook.active
            break
        except FileNotFoundError:
            print("File could not be found. Please try again.")
        except:
            print("Something went wrong with the input excel file...")


    try: 
        ind= find_col(sheet)
        if (ind == -1):
            ind = int(input("Please enter the number of the column containing the values: "))
    except:
        ind = int(input("Please enter the number of the column containing the values: "))
    
    vals = []   
    for i in range(1, sheet.max_row + 1):
        vals.append(str(sheet.cell(row = i, column = ind).value))
    return vals



def get_dois(lines):
  pattern = r'\b(10[.][0-9]{4,}(?:[.][0-9]+)*/(?:(?!["&\'<>])\S)+)\b'
  dois = []
  for txt in lines:
    dois.extend(re.findall(pattern, txt))
  return dois



def read_txt():
  lines = []
  while True:
    fname = input("Enter the name of the input text file: ")
    if ".txt" not in fname:
      fname += ".txt"
    try:
      with open(fname) as f_obj:
          lines = f_obj.readlines()
      break;
    except FileNotFoundError:
      print("Sorry, the file "+ fname + " does not exist. Please re-enter.")
    except:
      print("Something went wrong...")
  return lines
  


#-----------------------------------------------
  
choice = int(input("1 - text file\n2 - excel file\nSelect an option: "))
while (choice != 1 and choice != 2):
  choice = int(input("Please enter 1 or 2."))

if choice == 1: 
  dois = get_dois(read_txt())
else: 
  dois = get_dois(read_col())


#----------------------SELENIUM--------------------------

chrome_options = Options()
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument('--disable-dev-shm-usage')

driver = webdriver.Chrome(options=chrome_options)
driver.get("https://www.sciencedz.net/en/tools/doi2ris-doi-to-ris-converter.php")
try: 
    wait = WebDriverWait(driver, 5 * 60)#maximum wait time is 5 minutes
    # Wait until all elements are present on the page
    elements = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//*")))
except TimeoutException:
    # Handle the timeout exception
    print("Timeout occurred. The page did not fully load within the specified time.")
    driver.quit()
    sys.exit(1)


doi_box = driver.find_element(By.XPATH, '//*[@id="dois"]')
ris_box = driver.find_element(By.XPATH, '//*[@id="ris"]')
form = driver.find_element(By.TAG_NAME, 'form')
convert_btn = form.find_element(By.TAG_NAME, 'button')

ris = ''
out = ''
refnum = 1
for doi in dois: 
    doi_box.send_keys(doi)
    convert_btn.click()
    ref  = ''
    while (len(ref) < len(doi)):
        ref = ris_box.get_attribute("value")
        time.sleep(1)
    print("doi " + doi + " #" + str(refnum) + ": \n" + ref + "\n\n")
    ris += (ref + '\n')
    AU = ''
    PY = ''
    #extract author name
    i1 = ref.find('AU  -')
    i2 = ref.find(', ', i1)
    if i1 != -1 and i2 != -1:
        AU = ref[i1+6:i2]
    #find publication year
    i1 = ref.find('PY  -')
    if (i1 != -1):
        PY = ref[i1+6:i1+10]
    out += ('{' + AU + ', ' + PY + ' #' + str(refnum) + '}\n')
    refnum += 1
    doi_box.clear()
  
with open('output.txt', 'w') as file:
    file.write(out)
print("Done! Citations written to the file output.txt")

with open('output.ris', 'w', encoding='utf-8') as ris_file:
    ris_file.write(ris)
driver.close()

